
import pandas as pd
import openpyxl
from dash import Dash, dcc, html, Input, Output, dash_table
import plotly.express as px
import plotly.graph_objects as go

TELEMETRY_CSV = "moto3_goiania_telemetry.csv"
TASKS_CSV = "moto3_goiania_tasks.csv"
SETUP_CSV = "moto3_goiania_setup.csv"
ASPAR_XLSX = "Spec Domingo.xlsx"

df_telemetry = pd.read_csv(TELEMETRY_CSV)
df_tasks = pd.read_csv(TASKS_CSV)
df_setup = pd.read_csv(SETUP_CSV)

numeric_cols = [
    "presion_front_cold_bar", "presion_rear_cold_bar",
    "presion_front_hot_target_bar", "presion_rear_hot_target_bar",
    "track_temp_c", "air_temp_c", "humidity_pct", "wind_kmh",
    "lap_time_s", "sector_1_s", "sector_2_s", "sector_3_s",
    "velocidad_punta_kmh", "temp_neumatico_right_c",
    "temp_neumatico_center_c", "temp_neumatico_left_c",
    "anti_squat_pct", "wheelbase_delta_mm", "rake_delta_deg",
    "swingarm_pivot_delta_mm", "traction_control_lvl",
    "engine_brake_lvl", "anti_wheelie_lvl"
]
for col in numeric_cols:
    if col in df_telemetry.columns:
        df_telemetry[col] = pd.to_numeric(df_telemetry[col], errors="coerce")

roles = sorted(df_tasks["rol"].dropna().unique().tolist())
sesiones = ["FP1", "Practice", "FP2", "Q2", "Race"]
sesiones_disponibles = [s for s in sesiones if s in df_telemetry["sesion"].unique().tolist()]
if not sesiones_disponibles:
    sesiones_disponibles = sorted(df_telemetry["sesion"].dropna().unique().tolist())

def load_aspar_template(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    setting_headers = []
    for col in range(3, ws.max_column + 1):
        value = ws.cell(4, col).value
        if value is not None:
            setting_headers.append(str(value).strip())

    current_section = None
    rows = []
    known_sections = {"TYRES", "FORK", "SHOCK", "GEOMETRY", "ENGINE", "EXT CONDITION"}

    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue

        label_str = str(label).strip()
        label_upper = label_str.upper()
        cleaned_upper = label_upper.replace(" ", "")

        if label_upper in known_sections or cleaned_upper in {s.replace(" ", "") for s in known_sections}:
            current_section = label_str.strip()
            continue

        if label_upper in {"BIKE", "NOTES:"}:
            continue

        row = {"section": current_section if current_section else "GENERAL", "parameter": label_str}
        has_any_value = False

        for idx, col in enumerate(range(3, 3 + len(setting_headers))):
            header = setting_headers[idx]
            val = ws.cell(r, col).value
            row[header] = "" if val is None else str(val)
            if val not in (None, ""):
                has_any_value = True

        row["has_values"] = "Yes" if has_any_value else "No"
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, setting_headers

df_aspar, aspar_settings = load_aspar_template(ASPAR_XLSX)

def get_setup_value(param_name, default="N/D"):
    row = df_setup[df_setup["parametro"] == param_name]
    if row.empty:
        return default
    return row.iloc[0]["valor"]

def fmt_num(value, decimals=2, suffix=""):
    if pd.isna(value):
        return f"N/D{suffix}"
    return f"{value:.{decimals}f}{suffix}"

def status_color(value, good_min=None, good_max=None, low_alert=None, high_alert=None):
    if pd.isna(value):
        return "#6b7280"
    if good_min is not None and good_max is not None and good_min <= value <= good_max:
        return "#16a34a"
    if low_alert is not None and value < low_alert:
        return "#dc2626"
    if high_alert is not None and value > high_alert:
        return "#f59e0b"
    return "#1f77b4"

def card(title, value, subtitle="", color="#1f77b4"):
    return html.Div(
        [
            html.Div(title, style={"fontSize": "14px", "color": "#6b7280", "marginBottom": "6px"}),
            html.Div(value, style={"fontSize": "28px", "fontWeight": "700", "color": color}),
            html.Div(subtitle, style={"fontSize": "12px", "color": "#9ca3af", "marginTop": "6px"})
        ],
        style={"backgroundColor": "white", "padding": "16px", "borderRadius": "14px",
               "boxShadow": "0 2px 10px rgba(0,0,0,0.06)", "minWidth": "180px", "flex": "1"}
    )

def make_kanban_column(title, tasks, bg):
    children = [html.H4(title, style={"marginBottom": "12px"})]
    if len(tasks) == 0:
        children.append(html.Div("Sin tareas", style={"color": "#6b7280"}))
    else:
        for task in tasks:
            children.append(
                html.Div(task, style={"backgroundColor": "white", "padding": "10px", "borderRadius": "10px",
                                      "marginBottom": "10px", "boxShadow": "0 1px 4px rgba(0,0,0,0.05)",
                                      "fontSize": "13px"})
            )
    return html.Div(children, style={"backgroundColor": bg, "padding": "14px", "borderRadius": "14px",
                                     "width": "32%", "verticalAlign": "top", "minHeight": "260px"})

def build_setup_summary(dff):
    wheelbase = get_setup_value("wheelbase_delta_mm", dff["wheelbase_delta_mm"].iloc[0] if not dff.empty else "N/D")
    rake = get_setup_value("rake_delta_deg", dff["rake_delta_deg"].iloc[0] if not dff.empty else "N/D")
    anti_min = get_setup_value("anti_squat_target_min_pct", "108")
    anti_max = get_setup_value("anti_squat_target_max_pct", "112")
    swingarm = get_setup_value("swingarm_pivot_delta_mm", dff["swingarm_pivot_delta_mm"].iloc[0] if not dff.empty else "N/D")
    straight = get_setup_value("main_straight_m", "994")
    curves_right = get_setup_value("curves_right", "9")
    curves_left = get_setup_value("curves_left", "5")

    track_temp = dff["track_temp_c"].mean() if "track_temp_c" in dff.columns and not dff.empty else None
    air_temp = dff["air_temp_c"].mean() if "air_temp_c" in dff.columns and not dff.empty else None
    humidity = dff["humidity_pct"].mean() if "humidity_pct" in dff.columns and not dff.empty else None

    return html.Div([
        html.P(f"Wheelbase: +{wheelbase} mm"),
        html.P(f"Rake: {rake}°"),
        html.P(f"Swingarm Pivot: +{swingarm} mm"),
        html.P(f"Anti-squat objetivo: {anti_min}%–{anti_max}%"),
        html.P(f"Recta principal: {straight} m"),
        html.P(f"Asimetría del trazado: {curves_right} derechas / {curves_left} izquierdas"),
        html.P(f"Condiciones medias de la sesión: Track {fmt_num(track_temp, 1, ' °C')} | Aire {fmt_num(air_temp, 1, ' °C')} | Humedad {fmt_num(humidity, 1, '%')}")
    ])

app = Dash(__name__)
app.title = "Dashboard Moto3"

default_role = "Ingeniero de Pista" if "Ingeniero de Pista" in roles else roles[0]
default_session = "Practice" if "Practice" in sesiones_disponibles else sesiones_disponibles[0]
default_aspar_section = df_aspar["section"].dropna().unique().tolist()[0] if not df_aspar.empty else "TYRES"

goiania_layout = html.Div([
    html.Div([
        html.Div([
            html.Label("Selecciona rol"),
            dcc.Dropdown(id="rol-dropdown", options=[{"label": r, "value": r} for r in roles], value=default_role, clearable=False)
        ], style={"width": "32%", "display": "inline-block", "marginRight": "2%"}),
        html.Div([
            html.Label("Selecciona sesión"),
            dcc.Dropdown(id="sesion-dropdown", options=[{"label": s, "value": s} for s in sesiones_disponibles], value=default_session, clearable=False)
        ], style={"width": "32%", "display": "inline-block"})
    ], style={"marginBottom": "24px"}),

    html.Div(id="cards-row", style={"display": "flex", "gap": "14px", "marginBottom": "24px"}),

    html.Div([
        dcc.Graph(id="lap-chart", style={"width": "49%", "display": "inline-block"}),
        dcc.Graph(id="sector-chart", style={"width": "49%", "display": "inline-block", "marginLeft": "2%"})
    ]),

    html.Div([
        dcc.Graph(id="thermal-chart", style={"width": "49%", "display": "inline-block"}),
        dcc.Graph(id="pressure-chart", style={"width": "49%", "display": "inline-block", "marginLeft": "2%"})
    ]),

    html.Div([
        dcc.Graph(id="maps-chart", style={"width": "49%", "display": "inline-block"}),
        dcc.Graph(id="compound-chart", style={"width": "49%", "display": "inline-block", "marginLeft": "2%"})
    ]),

    html.H3("Kanban operativo por rol"),
    html.Div(id="kanban-row", style={"display": "flex", "gap": "2%", "marginBottom": "26px"}),

    html.H3("Resumen del setup"),
    html.Div(id="setup-summary", style={"backgroundColor": "white", "padding": "18px", "borderRadius": "14px", "boxShadow": "0 2px 10px rgba(0,0,0,0.06)"}),

    html.H3("Lectura táctica por rol", style={"marginTop": "24px"}),
    html.Div(id="role-insight", style={"backgroundColor": "white", "padding": "18px", "borderRadius": "14px", "boxShadow": "0 2px 10px rgba(0,0,0,0.06)"})
])

aspar_layout = html.Div([
    html.Div(id="aspar-cards-row", style={"display": "flex", "gap": "14px", "marginBottom": "24px"}),

    html.Div([
        html.Div([
            html.Label("Sección"),
            dcc.Dropdown(id="aspar-section-dropdown",
                         options=[{"label": s, "value": s} for s in df_aspar["section"].dropna().unique().tolist()],
                         value=default_aspar_section, clearable=False)
        ], style={"width": "32%", "display": "inline-block", "marginRight": "2%"}),
        html.Div([
            html.Label("Columna de setting a destacar"),
            dcc.Dropdown(id="aspar-setting-dropdown",
                         options=[{"label": s, "value": s} for s in aspar_settings],
                         value=aspar_settings[0] if aspar_settings else None, clearable=False)
        ], style={"width": "32%", "display": "inline-block"})
    ], style={"marginBottom": "20px"}),

    html.Div([
        dcc.Graph(id="aspar-section-chart", style={"width": "49%", "display": "inline-block"}),
        dcc.Graph(id="aspar-completeness-chart", style={"width": "49%", "display": "inline-block", "marginLeft": "2%"})
    ]),

    html.H3("Matriz de setup Aspar"),
    dash_table.DataTable(
        id="aspar-table",
        columns=[{"name": "Section", "id": "section"}, {"name": "Parameter", "id": "parameter"}] + [{"name": s, "id": s} for s in aspar_settings],
        data=df_aspar.to_dict("records"),
        page_size=15,
        sort_action="native",
        filter_action="native",
        style_table={"overflowX": "auto"},
        style_cell={"textAlign": "left", "fontFamily": "Arial", "fontSize": "13px", "padding": "8px"},
        style_header={"backgroundColor": "#111827", "color": "white", "fontWeight": "bold"},
        style_data_conditional=[{"if": {"filter_query": "{has_values} = 'No'"}, "backgroundColor": "#f9fafb", "color": "#6b7280"}]
    ),

    html.H3("Lectura de la prueba Aspar", style={"marginTop": "24px"}),
    html.Div(id="aspar-insight", style={"backgroundColor": "white", "padding": "18px", "borderRadius": "14px", "boxShadow": "0 2px 10px rgba(0,0,0,0.06)"})
])

app.layout = html.Div([
    html.H1("Dashboard Moto3 — Goiânia + Aspar"),
    html.P("Panel principal de Goiânia y nueva pestaña para la prueba Aspar basada en el Excel recibido."),
    dcc.Tabs(
        id="main-tabs",
        value="tab-goiania",
        children=[
            dcc.Tab(label="Goiânia 2026", value="tab-goiania", children=goiania_layout),
            dcc.Tab(label="Prueba Aspar", value="tab-aspar", children=aspar_layout),
        ]
    )
], style={"fontFamily": "Arial, sans-serif", "padding": "24px", "backgroundColor": "#f6f8fb"})

@app.callback(
    Output("cards-row", "children"),
    Output("lap-chart", "figure"),
    Output("sector-chart", "figure"),
    Output("thermal-chart", "figure"),
    Output("pressure-chart", "figure"),
    Output("maps-chart", "figure"),
    Output("compound-chart", "figure"),
    Output("kanban-row", "children"),
    Output("setup-summary", "children"),
    Output("role-insight", "children"),
    Input("rol-dropdown", "value"),
    Input("sesion-dropdown", "value")
)
def update_goiania(rol, sesion):
    dff = df_telemetry[df_telemetry["sesion"] == sesion].copy()
    tff = df_tasks[(df_tasks["sesion"] == sesion) & (df_tasks["rol"] == rol)].copy()

    best_lap = dff["lap_time_s"].min() if "lap_time_s" in dff.columns else None
    vmax = dff["velocidad_punta_kmh"].max() if "velocidad_punta_kmh" in dff.columns else None
    temp_right = dff["temp_neumatico_right_c"].mean() if "temp_neumatico_right_c" in dff.columns else None
    anti_squat = dff["anti_squat_pct"].mean() if "anti_squat_pct" in dff.columns else None
    p_rear = dff["presion_rear_hot_target_bar"].mean() if "presion_rear_hot_target_bar" in dff.columns else None

    cards = [
        card("Mejor vuelta", fmt_num(best_lap, 2, " s"), f"Sesión: {sesion}"),
        card("Velocidad punta", fmt_num(vmax, 0, " km/h"), f"Recta principal: {get_setup_value('main_straight_m', '994')} m"),
        card("Temp. flanco derecho", fmt_num(temp_right, 1, " °C"), "Control térmico neumático", color=status_color(temp_right, high_alert=95)),
        card("Anti-squat", fmt_num(anti_squat, 0, " %"), "Objetivo 108–112%", color=status_color(anti_squat, good_min=108, good_max=112)),
        card("Presión trasera hot", fmt_num(p_rear, 2, " bar"), "Mínimo legal Race: 1.65", color=status_color(p_rear, low_alert=1.65)),
    ]

    fig_lap = px.line(dff, x="vuelta", y="lap_time_s", markers=True, color="run", title=f"Evolución del tiempo por vuelta — {sesion}") if not dff.empty else go.Figure()
    fig_lap.update_layout(template="plotly_white", yaxis_title="Tiempo (s)", xaxis_title="Vuelta")

    sectors_df = dff.melt(id_vars=["vuelta", "run"], value_vars=["sector_1_s", "sector_2_s", "sector_3_s"],
                          var_name="sector", value_name="tiempo") if not dff.empty else pd.DataFrame(columns=["vuelta", "run", "sector", "tiempo"])
    fig_sector = px.bar(sectors_df, x="vuelta", y="tiempo", color="sector", barmode="group", title=f"Sectores por vuelta — {sesion}") if not sectors_df.empty else go.Figure()
    fig_sector.update_layout(template="plotly_white", yaxis_title="Tiempo (s)", xaxis_title="Vuelta")

    fig_thermal = go.Figure()
    if not dff.empty:
        fig_thermal.add_trace(go.Scatter(x=dff["vuelta"], y=dff["temp_neumatico_right_c"], mode="lines+markers", name="Flanco derecho"))
        fig_thermal.add_trace(go.Scatter(x=dff["vuelta"], y=dff["temp_neumatico_center_c"], mode="lines+markers", name="Centro"))
        fig_thermal.add_trace(go.Scatter(x=dff["vuelta"], y=dff["temp_neumatico_left_c"], mode="lines+markers", name="Flanco izquierdo"))
    fig_thermal.update_layout(title=f"Gradiente térmico del neumático — {sesion}", template="plotly_white", xaxis_title="Vuelta", yaxis_title="Temperatura (°C)")

    fig_pressure = go.Figure()
    if not dff.empty:
        fig_pressure.add_trace(go.Bar(x=dff["vuelta"], y=dff["presion_front_hot_target_bar"], name="Delantera hot"))
        fig_pressure.add_trace(go.Bar(x=dff["vuelta"], y=dff["presion_rear_hot_target_bar"], name="Trasera hot"))
        fig_pressure.add_hline(y=1.65, line_dash="dash", annotation_text="Mínimo trasero")
    fig_pressure.update_layout(title=f"Presiones dinámicas objetivo — {sesion}", template="plotly_white", xaxis_title="Vuelta", yaxis_title="Presión (bar)", barmode="group")

    fig_maps = go.Figure()
    if not dff.empty:
        fig_maps.add_trace(go.Scatter(x=dff["vuelta"], y=dff["traction_control_lvl"], mode="lines+markers", name="TC"))
        fig_maps.add_trace(go.Scatter(x=dff["vuelta"], y=dff["engine_brake_lvl"], mode="lines+markers", name="EBC"))
        fig_maps.add_trace(go.Scatter(x=dff["vuelta"], y=dff["anti_wheelie_lvl"], mode="lines+markers", name="AWC"))
    fig_maps.update_layout(title=f"Mapas electrónicos por vuelta — {sesion}", template="plotly_white", xaxis_title="Vuelta", yaxis_title="Nivel")

    compound_df = dff.groupby(["run", "neumatico_front", "neumatico_rear"], as_index=False).size().rename(columns={"size": "conteo"}) if not dff.empty else pd.DataFrame()
    fig_compound = px.bar(compound_df, x="run", y="conteo", color="neumatico_rear", pattern_shape="neumatico_front",
                          title=f"Compuestos por run — {sesion}", hover_data=["neumatico_front", "neumatico_rear"]) if not compound_df.empty else go.Figure()
    fig_compound.update_layout(template="plotly_white", xaxis_title="Run", yaxis_title="Número de registros")

    todo_tasks = tff[tff["estado"] == "Todo"]["tarea"].tolist()
    progress_tasks = tff[tff["estado"] == "In Progress"]["tarea"].tolist()
    done_tasks = tff[tff["estado"] == "Done"]["tarea"].tolist()
    kanban = [
        make_kanban_column("Todo", todo_tasks, "#eef2ff"),
        make_kanban_column("In Progress", progress_tasks, "#fff7ed"),
        make_kanban_column("Done", done_tasks, "#ecfdf5"),
    ]

    summary = build_setup_summary(dff)

    insights = {
        "Piloto": "Concéntrate en la estabilidad en T1, conservar el flanco derecho del trasero y modular gas en las primeras vueltas para llegar con tracción al final.",
        "Ingeniero de Pista": "Cruza sectores, velocidad punta, presiones y anti-squat para decidir si el setup mantiene el equilibrio entre estabilidad en recta y agilidad en el mixto.",
        "Telemétrico": "Prioriza el cruce entre temperatura de pista, presiones dinámicas, temperatura L/C/R del neumático y mapas electrónicos para validar la correlación con el modelo base.",
        "Jefe de Mecánicos": "Tu foco es operativo: swap de neumáticos, chequeo de torque, estado de suspensiones y protocolo sin errores entre runs.",
        "Técnico de Neumáticos": "Vigila especialmente el SC1 trasero y el flanco derecho. La clave es mantener presión legal y evitar sobretemperatura o degradación prematura."
    }
    role_insight = html.P(insights.get(rol, "Sin insight disponible para este rol."))

    return cards, fig_lap, fig_sector, fig_thermal, fig_pressure, fig_maps, fig_compound, kanban, summary, role_insight

@app.callback(
    Output("aspar-cards-row", "children"),
    Output("aspar-section-chart", "figure"),
    Output("aspar-completeness-chart", "figure"),
    Output("aspar-insight", "children"),
    Input("aspar-section-dropdown", "value"),
    Input("aspar-setting-dropdown", "value")
)
def update_aspar(section, selected_setting):
    dfa = df_aspar.copy()
    section_df = dfa[dfa["section"] == section].copy()

    n_sections = dfa["section"].nunique()
    n_parameters = len(section_df)
    filled_count = 0
    if selected_setting and selected_setting in section_df.columns:
        filled_count = section_df[selected_setting].replace("", pd.NA).dropna().shape[0]

    cards = [
        card("Secciones detectadas", str(n_sections), "Bloques del Excel"),
        card("Parámetros en sección", str(n_parameters), f"Sección: {section}"),
        card("Valores cargados", str(filled_count), f"Columna: {selected_setting}" if selected_setting else "Sin columna"),
        card("Settings disponibles", str(len(aspar_settings)), "SETTING 1–6"),
    ]

    counts = dfa.groupby("section", as_index=False).size().rename(columns={"size": "parametros"})
    fig_section = px.bar(counts, x="section", y="parametros", title="Parámetros por sección — plantilla Aspar")
    fig_section.update_layout(template="plotly_white", xaxis_title="Sección", yaxis_title="Nº parámetros")

    completeness = []
    for s in aspar_settings:
        non_empty = dfa[s].replace("", pd.NA).dropna().shape[0] if s in dfa.columns else 0
        completeness.append({"setting": s, "filled_rows": non_empty})
    completeness_df = pd.DataFrame(completeness)
    fig_complete = px.bar(completeness_df, x="setting", y="filled_rows", title="Completitud por columna de setting")
    fig_complete.update_layout(template="plotly_white", xaxis_title="Setting", yaxis_title="Filas con valor")

    insight_text = (
        f"La pestaña Aspar se ha construido a partir de la plantilla Excel recibida. "
        f"Ahora mismo funciona como comparador de setup por bloques ({', '.join(sorted(dfa['section'].dropna().unique()))}). "
        f"En la sección '{section}' hay {n_parameters} parámetros definidos. "
        f"Para '{selected_setting}', se detectan {filled_count} celdas con datos. "
        f"Si completas el Excel con valores reales de pista, esta pestaña pasará automáticamente de plantilla a panel comparativo operativo."
    )
    return cards, fig_section, fig_complete, html.P(insight_text)

if __name__ == "__main__":
    app.run(debug=True)
